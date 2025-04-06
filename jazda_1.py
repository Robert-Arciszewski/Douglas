import os
import csv
import re
import requests
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def slugify(text: str, max_length=100) -> str:
    text = text.lower()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_]+", "-", text)
    return text.strip("-")[:max_length]

def download_image(url: str, path: Path):
    if not url:
        return ""
    try:
        img_data = requests.get(url).content
        with open(path, 'wb') as handler:
            handler.write(img_data)
        return path.name
    except Exception as e:
        print(f"❌ Błąd pobierania zdjęcia: {url} -> {e}")
        return ""

# Konfiguracja
product_ids = [
    "3001040600", "3001040571", "5010332032", "5010332058", "5010069337",
    "5010112000", "m001926262", "m002073097", "5011701010", "5011701009",
    "5011068006", "5011060002", "5010525364", "5010521719", "5011016004",
    "5010916009", "5009416099", "5010109075"
]

headers = {
    "Accept": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}

base_dir = Path().resolve()
csv_filename = base_dir / "1_douglas.csv"
xlsx_filename = base_dir / "1_douglas.xlsx"
images_root_dir = base_dir / "douglas_1"
images_root_dir.mkdir(exist_ok=True)
raw_dir = base_dir / "raw_responses"
raw_dir.mkdir(exist_ok=True)

# Stałe kolumny
base_fieldnames = [
    "Brand", "Product Name", "Product Family", "Variant Name", "Price", "Volume",
    "Description", "Application", "Ingredients", "Warnings", "Bullet Points",
    "Breadcrumbs", "Variant Image (Preview)", "Swatch Path", "Main Image", "Main Image (Gray)",
    "All Images", "Manufacturer Info"
]

all_rows = []
image_column_names = set()

for product_id in product_ids:
    url = f"https://www.douglas.pl/api/v2/products/{product_id}?fields=FULL"
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
    except Exception as e:
        print(f"❌ Błąd przy ID: {product_id} – {e}")
        continue

    product_data = response.json()
    variants = product_data.get("variantOptions", [])
    brand = product_data.get("brand", {}).get("name", "")
    brand_line = product_data.get("brandLine", {}).get("name", "")
    base_name = product_data.get("baseProductName", "")
    categories = product_data.get("categories", [])
    breadcrumb_str = " > ".join(f'{c.get("name", "")} ({c.get("url", "")})' for c in categories)

    brand_dir = images_root_dir / slugify(brand)
    brand_dir.mkdir(parents=True, exist_ok=True)

    for variant in variants:
        variant_name = variant.get("variantName", "")
        full_variant_name = f"{brand_line} {base_name} {variant_name}".strip()
        variant_slug = slugify(full_variant_name)
        variant_dir = brand_dir / variant_slug
        variant_dir.mkdir(parents=True, exist_ok=True)

        swatch_url = variant.get("previewImage", {}).get("url", "")
        swatch_path = variant_dir / f"{variant_slug}_swatch.jpg"
        swatch_file_name = download_image(swatch_url, swatch_path)
        swatch_path_rel = str(swatch_path.relative_to(base_dir)).replace("\\", "/") if swatch_file_name else ""

        images = variant.get("images", [])
        image_urls = []
        image_file_paths = []

        if images:
            for i, img in enumerate(images):
                img_url = img["url"]
                img_path = variant_dir / f"{variant_slug}_{i}.jpg"
                file_name = download_image(img_url, img_path)
                rel_path = str(img_path.relative_to(base_dir)).replace("\\", "/")
                image_urls.append(img_url)
                image_file_paths.append(rel_path)
                col_name = f"Image {i+1}"
                image_column_names.add(col_name)

            gray_url = f"{images[0]['url']}&grid=true&imPolicy=grayScaled"
            gray_path = variant_dir / f"{variant_slug}_gray.jpg"
            gray_file = download_image(gray_url, gray_path)
            gray_file_path = str(gray_path.relative_to(base_dir)).replace("\\", "/")
        else:
            gray_file_path = ""

        manuf = variant.get("manufacturerAddress", {})
        manuf_address = ", ".join(filter(None, [
            manuf.get("street", ""), manuf.get("postalCode", ""),
            manuf.get("city", ""), manuf.get("country", "")
        ]))
        manufacturer_info = ", ".join(filter(None, [
            manuf.get("company", ""), manuf_address, manuf.get("webContactInformation", "")
        ])) if manuf else ""

        row = {
            "Brand": brand,
            "Product Name": f"{brand_line} {base_name}".strip(),
            "Product Family": product_data.get("productFamily", {}).get("name", ""),
            "Variant Name": variant_name,
            "Price": variant.get("priceData", {}).get("formattedValue", ""),
            "Volume": f"{variant.get('numberContentUnits', '')} {variant.get('contentUnitOfBaseNumberContentUnits', '')}",
            "Description": product_data.get("description", ""),
            "Application": product_data.get("application", ""),
            "Ingredients": product_data.get("ingredients", ""),
            "Warnings": ", ".join(variant.get("safetyInformationCodes", [])),
            "Bullet Points": " | ".join(product_data.get("bulletPoints", [])),
            "Breadcrumbs": breadcrumb_str,
            "Variant Image (Preview)": swatch_url,
            "Swatch Path": swatch_path_rel,
            "Main Image": image_urls[0] if image_urls else "",
            "Main Image (Gray)": gray_file_path,
            "All Images": " | ".join(image_urls),
            "Manufacturer Info": manufacturer_info
        }

        for i, path in enumerate(image_file_paths):
            col = f"Image {i+1}"
            row[col] = path

        all_rows.append(row)

# Finalne kolumny
final_fieldnames = base_fieldnames + sorted(image_column_names)

# Zapis do CSV
with open(csv_filename, "w", newline='', encoding="utf-8") as csv_file:
    writer = csv.DictWriter(csv_file, fieldnames=final_fieldnames)
    writer.writeheader()
    writer.writerows(all_rows)

# Zapis do XLSX
print("✅ Zapisano CSV. Generuję XLSX z hiperlinkami...")

df = pd.read_csv(csv_filename, encoding="utf-8")
df.to_excel(xlsx_filename, index=False)

# Dodawanie hyperlinków w Excelu
wb = load_workbook(xlsx_filename)
ws = wb.active
link_font = Font(color="0000FF", underline="single")
col_map = {cell.value: cell.column for cell in ws[1] if cell.value}

for row in range(2, ws.max_row + 1):
    for col_name in col_map:
        if col_name.startswith("Image ") or "Path" in col_name or "Main Image" in col_name:
            cell = ws.cell(row=row, column=col_map[col_name])
            if cell.value and isinstance(cell.value, str) and not cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = link_font

wb.save(xlsx_filename)
print(f"✅ Zapisano także do XLSX: {xlsx_filename.name}")
